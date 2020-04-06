import datetime
import os

from openpyxl import load_workbook

from tkinter_message import show_error

BASE_DIR = os.path.dirname((os.path.abspath(__file__)))

datetime_types = [datetime.datetime, datetime.date, datetime.time, datetime.timedelta]

def init_sheet(path):
    try:
        wb = load_workbook(path)
        return wb.active
    except FileNotFoundError:
        show_error("File not found", "File at " + str(path) + "not found")


def get_headers(active_sheet):
    active_calls = []
    i = 1
    while True:
        if active_sheet.cell(1, i).value:
            active_calls.append(active_sheet.cell(1, i).value)
            i += 1
        else:
            break
    return active_calls


def get_values(active_sheet, column):
    active_cells = []
    i = 2
    while True:
        if active_sheet.cell(i, column).value:
            # TODO: Support formatting as in xlsx
            if type(active_sheet.cell(i, column).value) is datetime.datetime:
                active_sheet.cell(i, column).value = active_sheet.cell(i, column).value.strftime('%d-%m-%Y')
            active_cells.append(active_sheet.cell(i, column).value)
            i += 1
        else:
            break
    return active_cells


def sort_dicts(match_dict):
    max_count = 0
    for value in match_dict.values():
        if len(value) > max_count:
            max_count = len(value)
    doc_replace_dicts = []
    for i in range(max_count):
        dicts = {}
        for keys, value in match_dict.items():
            try:
                dicts.update({keys: value[i]})
            except IndexError:
                dicts.update({keys: ""})

        doc_replace_dicts.append(dicts)
    return doc_replace_dicts



